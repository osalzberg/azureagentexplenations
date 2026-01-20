"""
Flask Web Application for Azure Log Analytics KQL Queries

This app provides a web interface to execute KQL queries against
Azure Log Analytics workspaces using Azure credentials.
"""

import os
import json
import time
from datetime import timedelta
from flask import Flask, render_template, request, jsonify
from dotenv import load_dotenv
from openai import AzureOpenAI
from monitor_client import AzureMonitorAgent

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Default workspace ID (can be overridden in the UI)
DEFAULT_WORKSPACE_ID = os.getenv("AZURE_LOG_ANALYTICS_WORKSPACE_ID", "")

# Available AI Models configuration
AI_MODELS = {
    "gpt-4": {
        "name": "GPT-4",
        "deployment": "gpt-4",
        "endpoint": os.getenv("AZURE_OPENAI_ENDPOINT"),
        "key": os.getenv("AZURE_OPENAI_KEY")
    },
    "gpt-4.1-nano": {
        "name": "GPT-4.1 Nano",
        "deployment": "gpt-4.1-nano",
        "endpoint": os.getenv("AZURE_OPENAI_ENDPOINT_2"),
        "key": os.getenv("AZURE_OPENAI_KEY_2")
    },
    "gpt-5.2-chat": {
        "name": "GPT-5.2 Chat",
        "deployment": "gpt-5.2-chat",
        "endpoint": os.getenv("AZURE_OPENAI_ENDPOINT_2"),
        "key": os.getenv("AZURE_OPENAI_KEY_2")
    },
    "o4-mini": {
        "name": "O4 Mini",
        "deployment": "o4-mini",
        "endpoint": os.getenv("AZURE_OPENAI_ENDPOINT_2"),
        "key": os.getenv("AZURE_OPENAI_KEY_2")
    }
}

DEFAULT_MODEL = "gpt-4"

# Audience-specific weight configurations
AUDIENCE_WEIGHTS = {
    "developer": {
        "faithfulness": 0.25,
        "structure": 0.10,
        "clarity": 0.15,
        "analysisDepth": 0.20,
        "contextAccuracy": 0.15,
        "actionability": 0.10,
        "conciseness": 0.05
    },
    "sre": {
        "faithfulness": 0.25,
        "structure": 0.08,
        "clarity": 0.12,
        "analysisDepth": 0.15,
        "contextAccuracy": 0.12,
        "actionability": 0.20,  # Higher for SRE - need actionable steps
        "conciseness": 0.08
    },
    "analyst": {
        "faithfulness": 0.30,  # Critical for analysts
        "structure": 0.12,
        "clarity": 0.18,
        "analysisDepth": 0.25,  # Deep analysis important
        "contextAccuracy": 0.10,
        "actionability": 0.05,
        "conciseness": 0.00
    },
    "executive": {
        "faithfulness": 0.20,
        "structure": 0.15,  # Clear structure important
        "clarity": 0.25,    # Crystal clear for executives
        "analysisDepth": 0.15,
        "contextAccuracy": 0.05,
        "actionability": 0.15,
        "conciseness": 0.05
    }
}

def get_calibration_examples(audience):
    """Get audience-specific calibration examples for scoring."""
    examples = {
        "developer": """
**Example of Score 5 (Faithfulness):** "The query returned 247 failed requests with ResultCode 500, representing 12% of total requests. The top affected endpoint is /api/users with 89 failures."

**Example of Score 3 (Faithfulness):** "The query shows several failed requests. This might indicate a server issue that should be investigated."

**Example of Score 1 (Faithfulness):** "The high failure rate of 45% (note: actual data shows 12%) is likely caused by database connection issues (no database mentioned in query)."
""",
        "sre": """
**Example of Score 5 (Actionability):** "1. Check application logs for endpoint /api/users between 14:00-15:00 UTC. 2. Review recent deployments in that timeframe. 3. Examine database query performance for user lookup operations."

**Example of Score 3 (Actionability):** "You should investigate the failed requests and check if there are any patterns."

**Example of Score 1 (Actionability):** "The system seems to have some issues."
""",
        "analyst": """
**Example of Score 5 (Analysis Depth):** "The temporal pattern shows failures spiking at 14:23 UTC (89 failures in 5 minutes), then declining. This correlates with increased traffic from region West-US-2, suggesting a regional load issue rather than code defect."

**Example of Score 3 (Analysis Depth):** "There are 247 failures shown in the data, with most happening in the afternoon hours."

**Example of Score 1 (Analysis Depth):** "The query shows: 247 rows of failed requests."
""",
        "executive": """
**Example of Score 5 (Clarity):** "**Status:** Service degradation detected. **Impact:** 12% of user requests failed today. **Root Cause:** API endpoint overload. **Timeline:** Issue started 2:23 PM, resolved by 3:15 PM."

**Example of Score 3 (Clarity):** "The KQL query filtered Requests table for Success==false and got 247 results with various ResultCodes."

**Example of Score 1 (Clarity):** "By using the where clause to filter on the Success boolean and then aggregating with summarize by ResultCode, we can see the distribution of HTTP status codes..."
"""
    }
    return examples.get(audience, examples["developer"])

def normalize_judge_scores(all_judge_scores, dimensions):
    """
    Normalize scores across judges to account for bias.
    Uses z-score normalization per judge, then rescales to 1-5 range.
    """
    if len(all_judge_scores) < 2:
        return None  # Need at least 2 judges for normalization
    
    try:
        import statistics
        
        # Calculate each judge's mean and std across all dimensions
        judge_stats = {}
        for judge_data in all_judge_scores:
            model = judge_data["model"]
            scores = [judge_data["scores"].get(dim, 3) for dim in dimensions]
            judge_stats[model] = {
                "mean": statistics.mean(scores),
                "std": statistics.stdev(scores) if len(scores) > 1 and statistics.stdev(scores) > 0 else 1.0
            }
        
        # Normalize each judge's scores (z-score)
        normalized_by_judge = {}
        for judge_data in all_judge_scores:
            model = judge_data["model"]
            stats = judge_stats[model]
            normalized_by_judge[model] = {}
            
            for dim in dimensions:
                raw_score = judge_data["scores"].get(dim, 3)
                # Z-score: (x - mean) / std
                z_score = (raw_score - stats["mean"]) / stats["std"]
                normalized_by_judge[model][dim] = z_score
        
        # Calculate global mean and std across all normalized scores
        all_normalized_scores = []
        for dim in dimensions:
            for model in normalized_by_judge:
                all_normalized_scores.append(normalized_by_judge[model][dim])
        
        global_mean = statistics.mean(all_normalized_scores)
        global_std = statistics.stdev(all_normalized_scores) if len(all_normalized_scores) > 1 else 1.0
        
        # Rescale to 1-5 range and average across judges
        final_scores = {}
        for dim in dimensions:
            dim_scores = []
            for model in normalized_by_judge:
                z_score = normalized_by_judge[model][dim]
                # Rescale: mean=3, std≈1 in 1-5 range
                rescaled = 3 + z_score
                # Clamp to 1-5 range
                rescaled = max(1, min(5, rescaled))
                dim_scores.append(rescaled)
            
            final_scores[dim] = sum(dim_scores) / len(dim_scores)
        
        return final_scores
    
    except Exception as e:
        print(f"[NORMALIZE] Error normalizing scores: {e}")
        return None  # Fall back to raw averaging

def get_openai_client(model_id):
    """Get an OpenAI client for the specified model."""
    model_config = AI_MODELS.get(model_id)
    if not model_config:
        return None, None
    
    endpoint = model_config.get("endpoint")
    key = model_config.get("key")
    deployment = model_config.get("deployment")
    
    if not endpoint or not key:
        return None, None
    
    client = AzureOpenAI(
        azure_endpoint=endpoint,
        api_key=key,
        api_version="2025-01-01-preview",
        timeout=60.0  # 60 second timeout
    )
    return client, deployment


@app.route("/")
def index():
    """Render the main query interface."""
    return render_template("index.html", default_workspace_id=DEFAULT_WORKSPACE_ID)


@app.route("/api/models")
def get_models():
    """Return available AI models."""
    models = [
        {"id": model_id, "name": config["name"]}
        for model_id, config in AI_MODELS.items()
        if config.get("endpoint") and config.get("key")
    ]
    return jsonify({"models": models, "default": DEFAULT_MODEL})


@app.route("/api/query", methods=["POST"])
def execute_query():
    """Execute a KQL query and return results as JSON."""
    try:
        data = request.get_json()
        workspace_id = data.get("workspace_id", "").strip()
        kql_query = data.get("query", "").strip()
        timespan_hours = data.get("timespan_hours", 1)

        if not workspace_id:
            return jsonify({"error": "Workspace ID is required"}), 400

        if not kql_query:
            return jsonify({"error": "Query is required"}), 400

        # Create Azure Monitor agent with default credentials
        agent = AzureMonitorAgent()

        # Set up timespan
        timespan = timedelta(hours=int(timespan_hours))

        # Execute the query
        result = agent.query_log_analytics(
            workspace_id=workspace_id,
            kql_query=kql_query,
            timespan=timespan
        )

        if "error" in result:
            return jsonify({"error": result["error"]}), 400

        # Process tables for JSON response
        tables = result.get("tables", [])
        processed_tables = []
        total_rows = 0

        for table in tables:
            # Convert rows to list of dicts for easier frontend consumption
            columns = table.get("columns", [])
            rows = table.get("rows", [])
            
            # Convert rows to list format if needed
            processed_rows = []
            for row in rows:
                if hasattr(row, '__iter__') and not isinstance(row, (str, dict)):
                    processed_rows.append(list(row))
                else:
                    processed_rows.append(row)

            processed_tables.append({
                "name": table.get("name", "Result"),
                "columns": columns,
                "rows": processed_rows,
                "row_count": len(processed_rows)
            })
            total_rows += len(processed_rows)

        return jsonify({
            "success": True,
            "tables": processed_tables,
            "total_rows": total_rows
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/test-connection", methods=["POST"])
def test_connection():
    """Test connection to a workspace."""
    try:
        data = request.get_json()
        workspace_id = data.get("workspace_id", "").strip()

        if not workspace_id:
            return jsonify({"error": "Workspace ID is required"}), 400

        agent = AzureMonitorAgent()
        
        # Simple test query
        result = agent.query_log_analytics(
            workspace_id=workspace_id,
            kql_query="print 'Connection successful'",
            timespan=timedelta(hours=1)
        )

        if "error" in result:
            return jsonify({
                "success": False,
                "message": f"Connection failed: {result['error']}"
            })

        return jsonify({
            "success": True,
            "message": "Successfully connected to workspace"
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Connection failed: {str(e)}"
        })


@app.route("/api/benchmark/evaluate", methods=["POST"])
def evaluate_explanation():
    """Evaluate an explanation using multiple LLM judges."""
    try:
        data = request.get_json()
        explanation = data.get("explanation", "")
        test_case = data.get("testCase", {})
        target_audience = data.get("targetAudience", "developer")

        # Truncate explanation to prevent huge payloads (increased from 3000)
        max_explanation_len = 5000
        if len(explanation) > max_explanation_len:
            explanation = explanation[:max_explanation_len] + "... [truncated]"
        
        # Limit result data size (increased from 5 to 10 rows)
        results = test_case.get('results', {})
        if 'rows' in results and len(results['rows']) > 10:
            results = {**results, 'rows': results['rows'][:10]}
        results_str = json.dumps(results, default=str)[:1200]

        # Use multiple models as judges (including O4 Mini)
        judge_models = ["gpt-4", "gpt-5.2-chat", "gpt-4.1-nano", "o4-mini"]
        all_judge_scores = []
        judge_notes = []

        # Get audience-specific examples for calibration
        audience_examples = get_calibration_examples(target_audience)
        
        evaluation_prompt = f"""You are an expert evaluator for Azure Log Analytics explanations. Your goal is to provide accurate, calibrated scores that differentiate quality.

## Context
- Target Audience: {target_audience}
- KQL Query: {test_case.get('query', 'N/A')[:500]}
- Result Data: {results_str}

## Explanation to Evaluate:
{explanation}

## Calibration Examples:
{audience_examples}

## Scoring Rubric (1-5 scale):

**Use the full scale:**
- 5 = Exceptional, exemplary work
- 4 = Good, above average with minor issues
- 3 = Adequate, meets basic requirements
- 2 = Below average, significant issues
- 1 = Poor, fails to meet requirements

### 1. Faithfulness (No Hallucinations - CRITICAL)
- 5: Every claim directly verifiable from the data, no unsupported inferences
- 4: Very accurate, only trivial reasonable inferences
- 3: Mostly accurate, some minor unsupported but plausible claims
- 2: Contains several claims not supported by visible data
- 1: Significant hallucinations, false metrics, or fabricated insights

### 2. Structure (Organization)
- 5: Excellent structure with clear sections, logical flow, highly scannable
- 4: Well organized with good headings and structure
- 3: Basic structure present, could be clearer
- 2: Poor organization, difficult to follow
- 1: No structure, wall of text

### 3. Clarity (Appropriate for {target_audience})
- 5: Crystal clear for target audience, perfect terminology level
- 4: Clear and understandable with minimal jargon issues
- 3: Understandable but has some clarity issues
- 2: Confusing, wrong terminology level, or unexplained concepts
- 1: Very unclear, inappropriate for audience

### 4. Analysis Depth (Insights Beyond Numbers)
- 5: Deep insights, patterns, root causes, implications clearly explained
- 4: Good analysis with meaningful interpretation
- 3: Basic interpretation, some analysis but mostly descriptive
- 2: Minimal interpretation, mostly just describing data
- 1: No analysis, only restates raw numbers

### 5. Context Accuracy (Azure/KQL Knowledge)
- 5: Expert-level Azure knowledge, perfect terminology and concepts
- 4: Strong Azure understanding, correct interpretations
- 3: Basic but correct Azure/KQL understanding
- 2: Some misunderstandings of Azure concepts
- 1: Fundamental errors in Azure/KQL interpretation

### 6. Actionability (Useful Recommendations)
- 5: Specific, actionable steps directly tied to findings
- 4: Good relevant recommendations
- 3: Generic but reasonable recommendations
- 2: Vague or partially relevant suggestions
- 1: No actionable recommendations or irrelevant advice

### 7. Conciseness (Communication Efficiency)
- 5: Perfectly concise, every sentence adds value
- 4: Efficient with minimal redundancy
- 3: Reasonable length, some unnecessary content
- 2: Too verbose or missing important information
- 1: Extremely verbose/repetitive OR critically incomplete

Respond ONLY with valid JSON (no markdown):
{{
    "faithfulness": <score 1-5>,
    "structure": <score 1-5>,
    "clarity": <score 1-5>,
    "analysisDepth": <score 1-5>,
    "contextAccuracy": <score 1-5>,
    "actionability": <score 1-5>,
    "conciseness": <score 1-5>,
    "confidence": <1-5, your confidence in this evaluation>,
    "evaluatorNotes": "<specific observations: what worked well, what needs improvement>"
}}"""

        # Get evaluations from each judge model
        for judge_model in judge_models:
            try:
                openai_client, deployment = get_openai_client(judge_model)
                if not openai_client:
                    print(f"[MULTI-JUDGE] Skipping {judge_model}: not configured")
                    continue

                print(f"[MULTI-JUDGE] Getting evaluation from {judge_model}...")
                
                # Retry logic for empty responses
                max_retries = 3
                response_text = None
                
                for attempt in range(max_retries):
                    # Handle different model API requirements
                    if judge_model.startswith("o"):
                        # O-series models: no system message, no temperature, use max_completion_tokens
                        combined_prompt = f"You are an expert evaluator. Respond only with valid JSON. No markdown code blocks.\n\n{evaluation_prompt}"
                        response = openai_client.chat.completions.create(
                            model=deployment,
                            messages=[
                                {"role": "user", "content": combined_prompt}
                            ],
                            max_completion_tokens=1500
                        )
                    elif judge_model in ["gpt-5.2-chat"]:
                        # GPT-5.2: uses max_completion_tokens
                        response = openai_client.chat.completions.create(
                            model=deployment,
                            messages=[
                                {"role": "system", "content": "You are an expert evaluator. Respond only with valid JSON."},
                                {"role": "user", "content": evaluation_prompt}
                            ],
                            max_completion_tokens=800
                        )
                    else:
                        # Standard models: max_tokens + temperature
                        response = openai_client.chat.completions.create(
                            model=deployment,
                            messages=[
                                {"role": "system", "content": "You are an expert evaluator. Respond only with valid JSON."},
                                {"role": "user", "content": evaluation_prompt}
                            ],
                            max_tokens=800,
                            temperature=0.3
                        )

                    response_text = response.choices[0].message.content
                    if response_text:
                        break
                    print(f"[MULTI-JUDGE] {judge_model} returned empty response (attempt {attempt + 1}/{max_retries})")
                    time.sleep(1)  # Brief pause before retry
                
                if not response_text:
                    print(f"[MULTI-JUDGE] {judge_model} failed after {max_retries} retries")
                    continue
                    
                response_text = response_text.strip()
                print(f"[MULTI-JUDGE] {judge_model} raw response: {response_text[:300]}")
                
                # Parse JSON response
                if response_text.startswith('```'):
                    response_text = response_text.split('```')[1]
                    if response_text.startswith('json'):
                        response_text = response_text[4:]
                
                judge_scores = json.loads(response_text)
                all_judge_scores.append({
                    "model": judge_model,
                    "scores": judge_scores
                })
                
                if judge_scores.get("evaluatorNotes"):
                    judge_notes.append(f"**{AI_MODELS[judge_model]['name']}**: {judge_scores['evaluatorNotes']}")
                
                print(f"[MULTI-JUDGE] {judge_model} scores parsed successfully")
                
            except Exception as judge_err:
                print(f"[MULTI-JUDGE] Error from {judge_model}: {str(judge_err)}")
                continue

        # If no judges succeeded, return error
        if not all_judge_scores:
            return jsonify({"error": "All judge models failed"}), 500

        # Calculate statistics and normalized scores
        dimensions = ["faithfulness", "structure", "clarity", "analysisDepth", "contextAccuracy", "actionability", "conciseness"]
        
        # Calculate raw averages and statistics
        raw_averaged_scores = {}
        score_statistics = {}
        
        for dim in dimensions:
            dim_scores = [j["scores"].get(dim, 3) for j in all_judge_scores]
            raw_averaged_scores[dim] = sum(dim_scores) / len(dim_scores)
            
            # Calculate statistics for consensus checking
            import statistics
            score_statistics[dim] = {
                "mean": raw_averaged_scores[dim],
                "std": statistics.stdev(dim_scores) if len(dim_scores) > 1 else 0,
                "min": min(dim_scores),
                "max": max(dim_scores),
                "range": max(dim_scores) - min(dim_scores)
            }
        
        # Check for consensus issues (high disagreement)
        high_disagreement_dims = [
            dim for dim, stats in score_statistics.items() 
            if stats["std"] > 1.0 or stats["range"] > 2
        ]
        
        # Normalize scores (optional z-score normalization per judge)
        normalized_scores = normalize_judge_scores(all_judge_scores, dimensions)
        
        # Final averaged scores (using normalized if available, otherwise raw)
        averaged_scores = {}
        for dim in dimensions:
            if normalized_scores:
                averaged_scores[dim] = round(normalized_scores[dim], 2)
            else:
                averaged_scores[dim] = round(raw_averaged_scores[dim], 2)
        
        # Add metadata
        averaged_scores["evaluatorNotes"] = "\n\n".join(judge_notes) if judge_notes else "No detailed notes available"
        averaged_scores["judgeCount"] = len(all_judge_scores)
        averaged_scores["judges"] = [j["model"] for j in all_judge_scores]
        averaged_scores["consensus"] = {
            "highDisagreement": high_disagreement_dims,
            "statistics": score_statistics
        }
        
        # Calculate average confidence if available
        confidences = [j["scores"].get("confidence", 3) for j in all_judge_scores]
        averaged_scores["averageConfidence"] = round(sum(confidences) / len(confidences), 2)
        
        print(f"[MULTI-JUDGE] Final scores from {len(all_judge_scores)} judges: {averaged_scores}")
        if high_disagreement_dims:
            print(f"[CONSENSUS WARNING] High disagreement on: {high_disagreement_dims}")

        return jsonify({"scores": averaged_scores, "individualJudges": all_judge_scores})

    except Exception as e:
        print(f"[MULTI-JUDGE] Error: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/explain", methods=["POST"])
def explain_results():
    """Generate an AI explanation of query results."""
    import time
    start_time = time.time()
    
    try:
        data = request.get_json()
        query = data.get("query", "")
        tables = data.get("tables", [])
        total_rows = data.get("total_rows", 0)
        model_id = data.get("model", DEFAULT_MODEL)
        
        print(f"[EXPLAIN] Starting explanation with model={model_id}, total_rows={total_rows}")

        # Get the appropriate client for the selected model
        openai_client, deployment = get_openai_client(model_id)
        
        if not openai_client:
            return jsonify({"error": f"Model '{model_id}' not configured"}), 500

        print(f"[EXPLAIN] Client ready, preparing prompt... ({time.time() - start_time:.2f}s)")

        # Prepare a summary of results for the AI
        results_summary = []
        for table in tables:
            table_info = {
                "name": table.get("name", "Unknown"),
                "columns": table.get("columns", []),
                "row_count": table.get("row_count", 0),
                "sample_rows": table.get("rows", [])[:5]  # First 5 rows as sample
            }
            results_summary.append(table_info)

        prompt = f"""You are an Azure Log Analytics expert. Analyze the following KQL query and its results, then provide a clear, helpful explanation.

## KQL Query:
```kql
{query}
```

## Results Summary:
- Total rows returned: {total_rows}
- Tables: {json.dumps(results_summary, indent=2, default=str)}

## Instructions:
1. **Query Explanation**: Briefly explain what this KQL query does in plain language.
2. **Results Analysis**: Describe what the results show - patterns, notable values, or insights.
3. **Table/Column Context**: If you recognize standard Azure tables (like Heartbeat, AzureActivity, requests, exceptions, traces, etc.), explain what these tables typically contain and what the columns mean.
4. **Insights**: Highlight any interesting findings, potential issues, or recommendations based on the data.

Keep the explanation concise but informative. Use bullet points for clarity. If the results are empty, explain possible reasons why."""

        print(f"[EXPLAIN] Prompt ready, calling API... ({time.time() - start_time:.2f}s)")

        # O-series models (o4-mini, o1, etc.) don't support system messages or temperature
        if model_id.startswith("o"):
            # Combine system prompt into user message for o-series models
            combined_prompt = f"""You are an expert in Azure Log Analytics, KQL (Kusto Query Language), and Azure monitoring. Provide clear, actionable explanations.

{prompt}"""
            try:
                print(f"[EXPLAIN] Calling O-series model {model_id}... (this may take longer)")
                response = openai_client.chat.completions.create(
                    model=deployment,
                    messages=[
                        {"role": "user", "content": combined_prompt}
                    ],
                    max_completion_tokens=4000
                )
            except Exception as o_err:
                print(f"[O-SERIES ERROR] {model_id}: {str(o_err)}")
                raise
        # Use max_completion_tokens for newer chat models (gpt-5.2, etc.)
        elif model_id in ["gpt-5.2-chat"]:
            print(f"[EXPLAIN] Calling GPT-5.2 API... ({time.time() - start_time:.2f}s)")
            response = openai_client.chat.completions.create(
                model=deployment,
                messages=[
                    {"role": "system", "content": "You are an expert in Azure Log Analytics, KQL (Kusto Query Language), and Azure monitoring. Provide clear, actionable explanations."},
                    {"role": "user", "content": prompt}
                ],
                max_completion_tokens=1000
            )
        else:
            print(f"[EXPLAIN] Calling {model_id} API... ({time.time() - start_time:.2f}s)")
            response = openai_client.chat.completions.create(
                model=deployment,
                messages=[
                    {"role": "system", "content": "You are an expert in Azure Log Analytics, KQL (Kusto Query Language), and Azure monitoring. Provide clear, actionable explanations."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1000,
                temperature=0.7
            )

        explanation = response.choices[0].message.content
        
        print(f"[EXPLAIN] Complete! Total time: {time.time() - start_time:.2f}s")

        return jsonify({
            "success": True,
            "explanation": explanation,
            "model": model_id
        })

    except Exception as e:
        print(f"[EXPLAIN] Error after {time.time() - start_time:.2f}s: {str(e)}")
        return jsonify({"error": str(e)}), 500


# KQL example queries for quick access
KQL_EXAMPLES = {
    "heartbeat": {
        "name": "Heartbeat & Health",
        "queries": [
            {
                "name": "Agent Heartbeats",
                "query": "Heartbeat\n| where TimeGenerated > ago(1h)\n| summarize LastHeartbeat=max(TimeGenerated) by Computer, OSType\n| order by LastHeartbeat desc"
            },
            {
                "name": "Heartbeat Count by Computer",
                "query": "Heartbeat\n| where TimeGenerated > ago(24h)\n| summarize HeartbeatCount=count() by Computer\n| order by HeartbeatCount desc"
            },
            {
                "name": "Missing Heartbeats",
                "query": "Heartbeat\n| where TimeGenerated > ago(1d)\n| summarize LastHeartbeat=max(TimeGenerated) by Computer\n| where LastHeartbeat < ago(15m)\n| order by LastHeartbeat asc"
            }
        ]
    },
    "azureactivity": {
        "name": "Azure Activity",
        "queries": [
            {
                "name": "Recent Activity",
                "query": "AzureActivity\n| where TimeGenerated > ago(24h)\n| project TimeGenerated, OperationName, ActivityStatus, Caller, ResourceGroup\n| order by TimeGenerated desc\n| take 100"
            },
            {
                "name": "Failed Operations",
                "query": "AzureActivity\n| where TimeGenerated > ago(24h)\n| where ActivityStatus == 'Failed'\n| summarize FailedCount=count() by OperationName, ResourceGroup\n| order by FailedCount desc"
            },
            {
                "name": "Activity by Caller",
                "query": "AzureActivity\n| where TimeGenerated > ago(7d)\n| summarize OperationCount=count() by Caller\n| order by OperationCount desc\n| take 20"
            }
        ]
    },
    "performance": {
        "name": "Performance",
        "queries": [
            {
                "name": "CPU Usage",
                "query": "Perf\n| where TimeGenerated > ago(1h)\n| where ObjectName == 'Processor' and CounterName == '% Processor Time'\n| summarize AvgCPU=avg(CounterValue) by Computer, bin(TimeGenerated, 5m)\n| order by TimeGenerated desc"
            },
            {
                "name": "Memory Usage",
                "query": "Perf\n| where TimeGenerated > ago(1h)\n| where ObjectName == 'Memory' and CounterName == '% Used Memory'\n| summarize AvgMemory=avg(CounterValue) by Computer\n| order by AvgMemory desc"
            },
            {
                "name": "Disk Free Space",
                "query": "Perf\n| where TimeGenerated > ago(1h)\n| where ObjectName == 'LogicalDisk' and CounterName == '% Free Space'\n| summarize AvgFreeSpace=avg(CounterValue) by Computer, InstanceName\n| where AvgFreeSpace < 20\n| order by AvgFreeSpace asc"
            }
        ]
    },
    "security": {
        "name": "Security",
        "queries": [
            {
                "name": "Security Events",
                "query": "SecurityEvent\n| where TimeGenerated > ago(24h)\n| summarize EventCount=count() by EventID, Activity\n| order by EventCount desc\n| take 20"
            },
            {
                "name": "Failed Logons",
                "query": "SecurityEvent\n| where TimeGenerated > ago(24h)\n| where EventID == 4625\n| summarize FailedLogons=count() by TargetAccount, Computer\n| order by FailedLogons desc"
            },
            {
                "name": "Account Lockouts",
                "query": "SecurityEvent\n| where TimeGenerated > ago(7d)\n| where EventID == 4740\n| project TimeGenerated, TargetAccount, Computer\n| order by TimeGenerated desc"
            }
        ]
    },
    "syslog": {
        "name": "Syslog (Linux)",
        "queries": [
            {
                "name": "Recent Syslog",
                "query": "Syslog\n| where TimeGenerated > ago(1h)\n| project TimeGenerated, Computer, Facility, SeverityLevel, SyslogMessage\n| order by TimeGenerated desc\n| take 100"
            },
            {
                "name": "Errors by Facility",
                "query": "Syslog\n| where TimeGenerated > ago(24h)\n| where SeverityLevel in ('err', 'crit', 'alert', 'emerg')\n| summarize ErrorCount=count() by Facility, Computer\n| order by ErrorCount desc"
            }
        ]
    }
}


@app.route("/api/examples")
def get_examples():
    """Return KQL example queries."""
    return jsonify(KQL_EXAMPLES)

@app.route("/api/audience-weights")
def get_audience_weights():
    """Return audience-specific scoring weights."""
    return jsonify(AUDIENCE_WEIGHTS)


@app.route("/api/benchmark/upload-excel", methods=["POST"])
def upload_excel():
    """Parse an uploaded Excel file containing KQL queries."""
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "Please upload an Excel file (.xlsx or .xls)"}), 400
        
        # Load workbook
        wb = load_workbook(filename=BytesIO(file.read()))
        ws = wb.active
        
        queries = []
        headers = [cell.value for cell in ws[1]] if ws[1] else []
        
        # Find query column (look for 'query', 'kql', 'queries' in header)
        query_col = None
        name_col = None
        description_col = None
        
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if header_lower in ['query', 'kql', 'queries', 'kql query']:
                    query_col = i
                elif header_lower in ['name', 'title', 'query name']:
                    name_col = i
                elif header_lower in ['description', 'desc', 'notes']:
                    description_col = i
        
        # If no header found, assume first column is query
        if query_col is None:
            query_col = 0
            # Check if first row looks like a header
            first_cell = ws.cell(row=1, column=1).value
            if first_cell and str(first_cell).lower() in ['query', 'kql', 'queries', 'name']:
                start_row = 2
            else:
                start_row = 1
        else:
            start_row = 2
        
        # Extract queries
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row[query_col] and str(row[query_col]).strip():
                query_text = str(row[query_col]).strip()
                query_name = str(row[name_col]).strip() if name_col is not None and row[name_col] else f"Query {len(queries) + 1}"
                query_desc = str(row[description_col]).strip() if description_col is not None and row[description_col] else ""
                
                queries.append({
                    "name": query_name,
                    "query": query_text,
                    "description": query_desc
                })
        
        if not queries:
            return jsonify({"error": "No queries found in the Excel file. Make sure queries are in a column labeled 'Query' or 'KQL'."}), 400
        
        return jsonify({
            "success": True,
            "queries": queries,
            "count": len(queries)
        })
        
    except Exception as e:
        return jsonify({"error": f"Failed to parse Excel file: {str(e)}"}), 500


@app.route("/api/benchmark/export-excel", methods=["POST"])
def export_excel():
    """Export benchmark results to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
        
        data = request.json
        results = data.get('results', {})
        queries = data.get('queries', [])
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        winner_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary headers
        summary_headers = ["Model", "Weighted Score", "Faithfulness", "Structure", "Clarity", 
                         "Analysis Depth", "Context Accuracy", "Actionability", "Conciseness"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Summary data
        leaderboard = results.get('leaderboard', [])
        for row_idx, model_data in enumerate(leaderboard, 2):
            ws_summary.cell(row=row_idx, column=1, value=model_data.get('model', ''))
            ws_summary.cell(row=row_idx, column=2, value=round(model_data.get('weightedScore', 0), 2))
            ws_summary.cell(row=row_idx, column=3, value=round(model_data.get('scores', {}).get('faithfulness', 0), 2))
            ws_summary.cell(row=row_idx, column=4, value=round(model_data.get('scores', {}).get('structure', 0), 2))
            ws_summary.cell(row=row_idx, column=5, value=round(model_data.get('scores', {}).get('clarity', 0), 2))
            ws_summary.cell(row=row_idx, column=6, value=round(model_data.get('scores', {}).get('analysisDepth', 0), 2))
            ws_summary.cell(row=row_idx, column=7, value=round(model_data.get('scores', {}).get('contextAccuracy', 0), 2))
            ws_summary.cell(row=row_idx, column=8, value=round(model_data.get('scores', {}).get('actionability', 0), 2))
            ws_summary.cell(row=row_idx, column=9, value=round(model_data.get('scores', {}).get('conciseness', 0), 2))
            
            # Highlight winner
            if row_idx == 2:
                for col in range(1, 10):
                    ws_summary.cell(row=row_idx, column=col).fill = winner_fill
            
            for col in range(1, 10):
                ws_summary.cell(row=row_idx, column=col).border = thin_border
        
        # Auto-width columns
        for col in range(1, 10):
            ws_summary.column_dimensions[get_column_letter(col)].width = 15
        
        # Per-Query Results Sheet
        ws_queries = wb.create_sheet("Per-Query Results")
        
        query_headers = ["Query #", "Query Name", "Model", "Weighted Score", 
                        "Faithfulness", "Structure", "Clarity", "Analysis Depth", 
                        "Context Accuracy", "Actionability", "Conciseness"]
        for col, header in enumerate(query_headers, 1):
            cell = ws_queries.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        row_idx = 2
        per_query = results.get('perQuery', [])
        for q_idx, query_result in enumerate(per_query):
            query_name = queries[q_idx].get('name', f'Query {q_idx + 1}') if q_idx < len(queries) else f'Query {q_idx + 1}'
            for model, model_result in query_result.get('modelResults', {}).items():
                scores = model_result.get('scores', {})
                ws_queries.cell(row=row_idx, column=1, value=q_idx + 1)
                ws_queries.cell(row=row_idx, column=2, value=query_name)
                ws_queries.cell(row=row_idx, column=3, value=model)
                ws_queries.cell(row=row_idx, column=4, value=round(model_result.get('weightedScore', 0), 2))
                ws_queries.cell(row=row_idx, column=5, value=round(scores.get('faithfulness', 0), 2))
                ws_queries.cell(row=row_idx, column=6, value=round(scores.get('structure', 0), 2))
                ws_queries.cell(row=row_idx, column=7, value=round(scores.get('clarity', 0), 2))
                ws_queries.cell(row=row_idx, column=8, value=round(scores.get('analysisDepth', 0), 2))
                ws_queries.cell(row=row_idx, column=9, value=round(scores.get('contextAccuracy', 0), 2))
                ws_queries.cell(row=row_idx, column=10, value=round(scores.get('actionability', 0), 2))
                ws_queries.cell(row=row_idx, column=11, value=round(scores.get('conciseness', 0), 2))
                
                for col in range(1, 12):
                    ws_queries.cell(row=row_idx, column=col).border = thin_border
                row_idx += 1
        
        # Auto-width
        for col in range(1, 12):
            ws_queries.column_dimensions[get_column_letter(col)].width = 15
        ws_queries.column_dimensions['B'].width = 30
        
        # Queries Sheet
        ws_query_list = wb.create_sheet("Queries")
        query_list_headers = ["#", "Name", "Query"]
        for col, header in enumerate(query_list_headers, 1):
            cell = ws_query_list.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for q_idx, query in enumerate(queries, 2):
            ws_query_list.cell(row=q_idx, column=1, value=q_idx - 1)
            ws_query_list.cell(row=q_idx, column=2, value=query.get('name', ''))
            ws_query_list.cell(row=q_idx, column=3, value=query.get('query', ''))
        
        ws_query_list.column_dimensions['A'].width = 5
        ws_query_list.column_dimensions['B'].width = 30
        ws_query_list.column_dimensions['C'].width = 80
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='benchmark_results.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": f"Failed to export Excel: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)
